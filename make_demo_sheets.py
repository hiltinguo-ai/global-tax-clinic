"""Generate synthetic, workable Excel worksheets for the six clinic personas.

Run:  python3 make_demo_sheets.py [outdir]
Out:  <outdir>/<persona>_*.xlsx  (6 workbooks)

Figures match clinic/personas.py gold profiles. Totals are live formulas, so
the sheets recalculate when edited. All persons, entities, and numbers are
fictional (synthetic demo data for the Global Tax Clinic, Sundai 30 Aug 2026).
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(sys.argv[1] if len(sys.argv) > 1 else "demo_docs")
OUT.mkdir(exist_ok=True)

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="0F6C78")
NOTE_FILL = PatternFill("solid", fgColor="F6E5DA")
TOTAL_FONT = Font(name=ARIAL, bold=True)
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF")
THIN = Border(bottom=Side(style="thin", color="D3DAE0"))

DISCLAIMER = ("Synthetic demo data for the Global Tax Clinic (Sundai, 30 Aug 2026). "
              "All persons, entities, accounts, and figures are fictional.")


def sheet(wb, title, headers, rows, money_cols=(), total_row=None, widths=None):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = title
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center")
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=ARIAL)
            cell.border = THIN
            if c in money_cols and isinstance(v, (int, float)):
                cell.number_format = "#,##0"
            if c in money_cols and isinstance(v, str) and v.startswith("="):
                cell.number_format = "#,##0"
    if total_row:
        r = len(rows) + 2
        for c, v in enumerate(total_row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = TOTAL_FONT
            if c in money_cols:
                cell.number_format = "#,##0"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = (widths[c - 1] if widths else 22)
    return ws


def notes_sheet(wb, lines):
    ws = wb.create_sheet("Notes")
    ws.cell(row=1, column=1, value="Notes & legend").font = Font(name=ARIAL, bold=True, size=12)
    r = 3
    for line in lines:
        cell = ws.cell(row=r, column=1, value=line)
        cell.font = Font(name=ARIAL)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Legend: edit only the plain rows; bold rows are formula totals and recalculate automatically.").font = Font(name=ARIAL, italic=True)
    d = ws.cell(row=r + 2, column=1, value=DISCLAIMER)
    d.font = Font(name=ARIAL, size=8, color="8A939C")
    d.fill = NOTE_FILL
    ws.column_dimensions["A"].width = 110


# ============================ MEI ============================
wb = Workbook()
sheet(
    wb, "Foreign accounts",
    ["Institution / account", "Type", "Country", "Max balance USD (year)"],
    [
        ["HSBC bank account", "bank", "Hong Kong", 3200],
        ["Hang Seng bank account", "bank", "Hong Kong", 2800],
        ["HK brokerage account", "brokerage", "Hong Kong", 4100],
        ["MPF retirement account", "mpf", "Hong Kong", 1400],
        ["HK retail equity fund", "fund", "Hong Kong", 3500],
    ],
    money_cols=(4,),
    total_row=["Aggregate maximum", "", "", "=SUM(D2:D6)"],
    widths=[28, 12, 14, 22],
)
sheet(
    wb, "Income and gifts",
    ["Item", "Kind", "Country", "Amount USD (2026)"],
    [
        ["Shenzhen apartment - rent out", "rental", "China", 18000],
        ["RSU vesting - US employer (Meridian Cloud Inc.)", "rsu", "US", 0],
        ["Cash gift from parents in mainland China (foreign persons)", "gift", "China", 180000],
    ],
    money_cols=(4,),
    widths=[48, 10, 12, 20],
)
notes_sheet(wb, [
    "Mei Zhang - personal worksheet, tax year 2026.",
    "US person (US tax resident), filing single, California. About 40 days in mainland China this year.",
    "The HK retail equity fund may be a passive foreign investment company.",
])
wb.save(OUT / "mei_worksheet.xlsx")

# ============================ LUIS ============================
wb = Workbook()
sheet(
    wb, "Income",
    ["Source", "Form", "State", "Amount USD (2026)"],
    [
        ["Beacon Hill Bistro - wages (wait tables)", "W-2", "MA", 32000],
        ["Cash and charged tips", "tips", "MA", 14000],
        ["DoorDash - delivery (1099-NEC, gig)", "1099-NEC", "MA", 9000],
    ],
    money_cols=(4,),
    total_row=["Total income", "", "", "=SUM(D2:D4)"],
    widths=[40, 12, 8, 20],
)
sheet(
    wb, "Residency",
    ["Period", "State", "Detail"],
    [
        ["Jan 1 - Jul 15, 2026", "Massachusetts", "Boston apartment"],
        ["Jul 15 - Dec 31, 2026", "New Hampshire", "Moved to Nashua (part-year resident of both states)"],
    ],
    widths=[22, 16, 50],
)
notes_sheet(wb, [
    "Luis Ramirez - 2026 income worksheet. US person, head of household, one child (dependent).",
    "No foreign accounts. No gifts received.",
])
wb.save(OUT / "luis_worksheet.xlsx")

# ============================ SICHUAN GARDEN ============================
wb = Workbook()
rows = []
for i, mon in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 2):
    rows.append([mon, 31000, 4000, f"=B{i}*0.0625", f"=B{i}*0.0075"])
ws = sheet(
    wb, "Monthly sales",
    ["Month 2026", "Meals sales USD", "Merch sales USD", "MA sales tax 6.25%", "Local meals excise 0.75%"],
    rows,
    money_cols=(2, 3, 4, 5),
    total_row=["Total", "=SUM(B2:B13)", "=SUM(C2:C13)", "=SUM(D2:D13)", "=SUM(E2:E13)"],
    widths=[12, 18, 18, 20, 22],
)
sheet(
    wb, "Payroll",
    ["Role", "Headcount", "Annual payroll USD"],
    [
        ["Kitchen", 6, 84000],
        ["Waitstaff", 7, 77000],
        ["Manager", 1, 24000],
    ],
    money_cols=(3,),
    total_row=["Total employees / payroll", "=SUM(B2:B4)", "=SUM(C2:C4)"],
    widths=[26, 12, 20],
)
notes_sheet(wb, [
    "Sichuan Garden LLC is a Massachusetts single-member LLC restaurant (Brookline, MA). The company serves meals.",
    "Owner is a US person. Employees on payroll. Registered on MassTaxConnect for sales, meals, and withholding.",
])
wb.save(OUT / "sichuan_books.xlsx")

# ============================ NIMBUSFLOW ============================
wb = Workbook()
sheet(
    wb, "Revenue by state",
    ["Billing state", "Customers", "Q1", "Q2", "Q3", "Q4", "2025 revenue USD", "SaaS taxable?"],
    [
        ["Massachusetts sales", 14, 100000, 104000, 106000, 110000, "=SUM(C2:F2)", "yes (6.25%)"],
        ["New York sales", 8, 57000, 59000, 61000, 63000, "=SUM(C3:F3)", "yes"],
        ["California sales", 6, 43000, 44000, 46000, 47000, "=SUM(C4:F4)", "yes"],
        ["Texas sales", 5, 29000, 30000, 30000, 31000, "=SUM(C5:F5)", "exempt"],
        ["Illinois sales", 3, 21000, 22000, 23000, 24000, "=SUM(C6:F6)", "yes"],
        ["Other states", 14, 48000, 50000, 50000, 52000, "=SUM(C7:F7)", "review"],
    ],
    money_cols=(3, 4, 5, 6, 7),
    total_row=["Total revenue", "=SUM(B2:B7)", "=SUM(C2:C7)", "=SUM(D2:D7)", "=SUM(E2:E7)", "=SUM(F2:F7)", "=SUM(G2:G7)", ""],
    widths=[22, 11, 11, 11, 11, 11, 18, 14],
)
sheet(
    wb, "Payroll",
    ["Employee", "Work state", "Annual payroll USD"],
    [
        ["A. Osei (CEO)", "Massachusetts", 190000],
        ["R. Bhatt (CTO)", "Massachusetts", 180000],
        ["K. Doyle (Sales)", "Massachusetts", 120000],
    ],
    money_cols=(3,),
    total_row=["Total payroll", "", "=SUM(C2:C4)"],
    widths=[24, 16, 20],
)
notes_sheet(wb, [
    "NimbusFlow, Inc. is a Massachusetts C-corporation (incorporated 15 March 2023), SaaS, 12 Winter St, Boston.",
    "The company is a US person. Three employees, all in Massachusetts. No foreign accounts.",
    "Out-of-state receipts have not been reviewed for economic nexus.",
])
wb.save(OUT / "nimbus_revenue.xlsx")

# ============================ NORI ROBOTICS ============================
wb = Workbook()
sheet(
    wb, "Cap table",
    ["Holder", "Country", "Shares", "Ownership"],
    [
        ["Founder A (Chinese founder)", "China", 4000000, "=C2/$C$6"],
        ["Founder B (Chinese founder)", "China", 1800000, "=C3/$C$6"],
        ["Employee pool", "US", 1200000, "=C4/$C$6"],
        ["Seed investors", "US", 3000000, "=C5/$C$6"],
    ],
    money_cols=(3,),
    total_row=["Total shares", "", "=SUM(C2:C5)", ""],
    widths=[30, 10, 14, 12],
)
ws = wb["Cap table"]
for r in range(2, 6):
    ws.cell(row=r, column=4).number_format = "0.0%"
sheet(
    wb, "Intercompany",
    ["Quarter 2026", "Item", "Amount USD"],
    [
        ["Q1", "Cost-plus engineering services - Shanghai subsidiary", 92000],
        ["Q2", "Cost-plus engineering services - Shanghai subsidiary", 95000],
        ["Q3", "Cost-plus engineering services - Shanghai subsidiary", 96000],
        ["Q4", "Cost-plus engineering services - Shanghai subsidiary", 97000],
    ],
    money_cols=(3,),
    total_row=["Total", "reimbursements to 100% Shanghai subsidiary", "=SUM(C2:C5)"],
    widths=[14, 52, 16],
)
notes_sheet(wb, [
    "Nori Robotics Inc. is a Delaware C-corporation with operations in Massachusetts. The company is a US person.",
    "The founders are Chinese tax residents and own more than 25 percent. Wholly owned subsidiary in Shanghai, China.",
    "Unvested founder stock granted 1 August 2026 (83(b) election window open, 30-day deadline).",
])
wb.save(OUT / "nori_captable.xlsx")

# ============================ CHEN FAMILY TRUST ============================
wb = Workbook()
sheet(
    wb, "Holdings",
    ["Holding", "Type", "Shares held", "Shares outstanding", "Ownership"],
    [
        ["Northline Inc. (NYSE-listed, founded by settlor ~20 years ago)", "public company shares", 9600000, 120000000, "=C2/D2"],
        ["Money-market sweep (US)", "cash", "", "", ""],
    ],
    money_cols=(3, 4),
    widths=[52, 20, 14, 18, 12],
)
wb["Holdings"].cell(row=2, column=5).number_format = "0.0%"
sheet(
    wb, "Foreign accounts",
    ["Institution / account", "Type", "Country", "Max balance USD (year)"],
    [
        ["HSBC Hong Kong bank account", "bank", "Hong Kong", 250000],
        ["HK brokerage custody account", "brokerage", "Hong Kong", 1750000],
    ],
    money_cols=(4,),
    total_row=["Aggregate maximum", "", "", "=SUM(D2:D3)"],
    widths=[34, 12, 14, 22],
)
sheet(
    wb, "Dividends",
    ["Quarter 2026", "Dividend per share USD", "Shares", "Dividend received USD"],
    [
        ["Q1", 0.22, 9600000, "=B2*C2"],
        ["Q2", 0.22, 9600000, "=B3*C3"],
        ["Q3", 0.23, 9600000, "=B4*C4"],
        ["Q4", 0.23, 9600000, "=B5*C5"],
    ],
    money_cols=(3, 4),
    total_row=["Total dividends", "", "", "=SUM(D2:D5)"],
    widths=[14, 20, 14, 22],
)
notes_sheet(wb, [
    "The Chen Family Trust is a US family trust with Massachusetts situs. The trust is a US person.",
    "The settlor founded Northline Inc. and took it public on the NYSE about 20 years ago; the trust still owns 8 percent of the listed shares.",
    "Northline Inc. operates in the US, mainland China, Hong Kong, and the EU through local subsidiaries.",
    "Mr. Chen and his family moved to Hong Kong 10 years ago (2016); he is a US person and a Hong Kong resident.",
    "Dividends are paid to the trust each year and partly distributed to beneficiaries.",
])
wb.save(OUT / "chen_trust.xlsx")

print("done:", sorted(p.name for p in OUT.glob("*.xlsx")))
