from __future__ import annotations

import io
import zipfile
from pathlib import Path

from clinic.schemas import UploadedDoc

MAX_BYTES = 8 * 1024 * 1024
MAX_CHARS = 20_000
TEXT_SUFFIXES = {".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".html"}


def ingest_files(files: list[tuple[str, bytes]]) -> tuple[str, list[UploadedDoc]]:
    chunks: list[str] = []
    docs: list[UploadedDoc] = []
    for name, data in files:
        if not data or len(data) > MAX_BYTES:
            continue
        suffix = Path(name).suffix.lower()
        try:
            if suffix == ".zip":
                text, inner = _zip(name, data)
            elif suffix == ".pdf":
                text = _pdf(data)
                inner = [UploadedDoc(name=name, kind="pdf", characters=len(text), excerpt=text[:400])]
            elif suffix in {".xlsx", ".xlsm"}:
                text = _xlsx(data)
                inner = [UploadedDoc(name=name, kind="spreadsheet", characters=len(text), excerpt=text[:400])]
            elif suffix in TEXT_SUFFIXES:
                text = data.decode("utf-8", errors="replace")
                inner = [UploadedDoc(name=name, kind="text", characters=len(text), excerpt=text[:400])]
            else:
                continue
        except Exception:
            continue
        if text.strip():
            chunks.append(f"--- {name} ---\n{text[:MAX_CHARS]}")
            docs.extend(inner)
    return "\n\n".join(chunks), docs


def _xlsx(data: bytes) -> str:
    """Cell values as text, one line per row. Reads cached formula results."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return ""
    parts: list[str] = []
    for ws in wb.worksheets[:10]:
        parts.append(f"## {ws.title}")
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r > 300:
                break
            vals = []
            for v in row:
                if v is None or v == "":
                    continue
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                vals.append(str(v))
            if vals:
                parts.append(" ".join(vals))
    return "\n".join(parts)


def _pdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for page in reader.pages[:40]:
        pages.append(page.extract_text() or "")
    return "\n".join(pages)


def _zip(archive_name: str, data: bytes) -> tuple[str, list[UploadedDoc]]:
    docs: list[UploadedDoc] = []
    parts: list[str] = []
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        for info in zf.infolist()[:30]:
            if info.is_dir() or info.file_size > MAX_BYTES:
                continue
            name = Path(info.filename).name
            if ".." in Path(info.filename).parts:
                continue
            suffix = Path(name).suffix.lower()
            raw = zf.read(info)
            if suffix == ".pdf":
                text = _pdf(raw)
                kind = "pdf"
            elif suffix in {".xlsx", ".xlsm"}:
                text = _xlsx(raw)
                kind = "spreadsheet"
            elif suffix in TEXT_SUFFIXES:
                text = raw.decode("utf-8", errors="replace")
                kind = "text"
            else:
                continue
            if not text.strip():
                continue
            parts.append(f"--- {archive_name}/{name} ---\n{text[:MAX_CHARS]}")
            docs.append(UploadedDoc(name=f"{archive_name}/{name}", kind=kind, characters=len(text), excerpt=text[:400]))
    return "\n\n".join(parts), docs
